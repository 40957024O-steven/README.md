import pandas as pd
import tkinter as tk
import tkinter.ttk as ttk
from tkinter import scrolledtext,font
import time
from openpyxl import load_workbook
import os


# 创建一个临时的Tkinter窗口
root = tk.Tk()

# 获取屏幕的宽度和高度
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

# 关闭临时的Tkinter窗口
root.destroy()


button_pressed = False
root = tk.Tk()



root.geometry(str(root.winfo_screenwidth())+"x"+str(root.winfo_screenheight()))
root.title("升學機器人")


path = os.path.join(os.getcwd(), '專題/期末專案/專案/112校系分則.xlsx')
# print(path)
wb = load_workbook(path)
ws = wb['112校系分則v1']
rowmax = ws.max_row
# print(rowmax)


def Gors(gors):  # 判斷學校是公立還是私立
    if gors == "國立" or gors == '公立':
        for i in range(3, rowmax + 1):
            schoolallname = ws.cell(i, 3).value  # 學校名稱
            # print(schoolallname)
            gors = schoolallname[:2]  # 公私立
            # print(gors)
            if schoolallname != ws.cell(i - 1, 3).value:
                if gors == '國立':
                    text_area.insert(tk.END, schoolallname + "\n", "tag_name")
        text_area.insert(tk.END, '\n\n請問要查詢的學校?' + "\n", "tag_name")

    elif gors == "私立":
        for i in range(3, rowmax + 1):
            schoolallname = ws.cell(i, 3).value  # 學校名稱
            # print(schoolallname)
            gors = schoolallname[:2]  # 公私立
            # print(gors)
            if schoolallname != ws.cell(i - 1, 3).value:
                if gors != '國立':
                    text_area.insert(tk.END, schoolallname + "\n", "tag_name")
        text_area.insert(tk.END, '\n\n請問要查詢的學校?' + "\n", "tag_name")


def School(school):  # 查詢的學校名稱
    chack = 0
    for i in range(3, rowmax + 1):
        gp = ws.cell(i, 4).value.replace(' ', '')
        if school == str(ws.cell(i, 3).value).replace(' ', ''):
            text_area.insert(tk.END, gp + "\n", "tag_name")
            serch[0] = school
            chack += 1
    if chack == 0:
        text_area.insert(tk.END, "查無此學校\n", "tag_name")
    print(serch)
    text_area.insert(tk.END, '\n\n請問要查詢的系所?' + "\n", "tag_name")


def Sci(sci):  # 查詢系所
    chack = 0
    for i in range(3, rowmax + 1):
        if sci == str(ws.cell(i, 4).value).replace(' ', '') and serch[0] == str(ws.cell(i, 3).value).replace(' ', ''):
            serch[1] = sci
            det(serch[0], serch[1], i)
            # text_area.insert(tk.END,str(i)+serch[0]+serch[1]+"\n","tag_name")
            chack += 1
    if chack == 0:
        text_area.insert(tk.END, serch[0] + "沒有此系\n", "tag_name")
    print(serch)


serch = [[], []]  # 搜尋[[學校],[系所]]


def det(school, sci, lc):
    text_area.insert(tk.END, serch[0] + serch[1] + "：\n", "tag_name")
    if ws.cell(lc, 5).value != None or ws.cell(lc, 6).value != None or ws.cell(lc, 7).value != None or ws.cell(lc, 8).value != None or ws.cell(lc, 9).value != None or ws.cell(lc, 10).value != None:
        text_area.insert(tk.END, " 學測檢定標準\n", "tag_title")
        for i in range(5, 11):
            if ws.cell(lc, i).value != None:
                text_area.insert(tk.END, "  " + ws.cell(2, i).value + " ==> " + ws.cell(lc, i).value + "\n", "inf")
    if ws.cell(lc, 11).value != None:
        text_area.insert(tk.END, "\n 英聽檢定標準 ==> ", "tag_title")
        text_area.insert(tk.END, ws.cell(lc, 11).value + "\n", "inf")
    text_area.insert(tk.END, "\n 採計科目及加權\n", "tag_title")
    for i in range(12, 26):
        if ws.cell(lc, i).value != None:
            text_area.insert(tk.END, "  " + ws.cell(2, i).value + " ==> " + ws.cell(lc, i).value + "\n", "inf")
    if ws.cell(lc, 26).value != None:
        text_area.insert(tk.END, "\n 術科類別 ==> ", "tag_title")
        text_area.insert(tk.END, ws.cell(lc, 26).value, "inf")
    text_area.insert(tk.END, "\n 同分參酌順序 ==> ", "tag_title")
    i = 27
    while ws.cell(lc, i).value != None:
        text_area.insert(tk.END, ws.cell(lc, i).value, "inf")
        if ws.cell(lc, i + 1).value != None and i != 31:
            text_area.insert(tk.END, "=>", "inf")
        i += 1
        if i == 32:
            break
    text_area.insert(tk.END,'\n')


def Conversation(conversation):
    if conversation == '國立' or conversation == '公立' or conversation == '私立':
        Gors(conversation)
        # text_area.insert(tk.END,'公私立查詢'+"\n","tag_name")

    elif conversation[-2:] == '大學':
        School(conversation)
        # text_area.insert(tk.END,'學校名稱查詢'+"\n","tag_name")
    elif conversation[-1] == '系' or conversation[-1] == ')' or conversation[-1] == "組"or conversation[-1] == "班"or conversation[-1] == "程":
        Sci(conversation)
        # text_area.insert(tk.END,'系所查詢'+"\n","tag_name")
    else:
        text_area.insert(tk.END, '輸入錯誤,請重輸入' + "\n", "tag_title")
    print(conversation)

custom_font_25 = font.Font(family="Arial", size=25)
custom_font_20 = font.Font(family="Arial", size=20)
custom_font_17 = font.Font(family="Arial", size=17)
# 輸出螢幕
text_area = scrolledtext.ScrolledText(root, width=screen_width, height=screen_height*0.9)
text_area.tag_config("tag_name", foreground="green",font=custom_font_17)
text_area.tag_config("tag_title", foreground="red",font=custom_font_20)
text_area.tag_config("inf", foreground="black",font=custom_font_20)
text_area.tag_config("cf", foreground="red",font=custom_font_25)
text_area.insert(tk.END, "溫馨提醒：請開啟全螢幕\n", "cf")
text_area.insert(tk.END, "機器人：\n請輸入以下指令進行互動\n您所要查詢的大學是國立(公立)OR私立?\n", "tag_name")

text_area.pack()


def handle_button_press():
    # 获取对话框的文本
    t = enterwindow.get("1.0", "end").replace('\n', '').replace(' ', '')
    # print(t)
    # 清空对话框的文本
    enterwindow.delete("1.0", tk.END)
    # 在输出框中添加文本
    text_area.insert(tk.END, "\n您：\n{}\n\n".format(t))
    # print(t)
    text_area.insert(tk.END, "機器人：\n", "tag_name")
    # print(t)
    Conversation(t)



# 對話框
enter_width = screen_width-525
enter_height = screen_height-320
but_width = enter_width+440
but_height = screen_height-320
enterwindow = tk.Text(root, width=60, height=15,bd=4,relief='solid')
enterwindow.place(x=enter_width,y=enter_height)
enterwindow.configure(bg='floralwhite')

# 輸入按鈕
bt = tk.Button(root, text='輸入', width=8, height=4, command=handle_button_press)
bt.place(x=but_width,y=but_height)

root.mainloop()

