import os
from openpyxl import load_workbook

path = os.path.join(os.getcwd(),'期末專案/112校系分則.xlsx')
# print(path)
wb = load_workbook(path)
ws =  wb['112校系分則v1']
rowmax = ws.max_row
school = '國立臺灣大學'

for i in range(3,30):
    gp = ws.cell(i,4).value
    if school == ws.cell(i,3).value.replace(' ',''):
        print(gp)
    else:
        print(ws.cell(i,3).value+ str(len(school)))