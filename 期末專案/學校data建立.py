from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from openpyxl import Workbook

chrome_options = Options()
path = "C:/Users/steve/Downloads/chromedriver_win32/chromedriver.exe" # chromedriver位置
service = Service(path)

driver = webdriver.Chrome(service=service, options=chrome_options)



driver.get("https://university-tw.ldkrsi.men/uac/") # 目標頁面

# 學校 /html/body/main/section[3]/ul/li[1]/a ==> 其中section[3] = 私立 section[2] = 公立
# 程式進行到一辦可能會卡住，原因不明，但是讓程式重中間繼續執行就好

for t in range(17,32):
    print(t)
    element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "/html/body/main/section[3]/ul/li[1]/a"))) #系統搜尋到某些東西時才繼續進行

    driver.find_element(By.XPATH,"/html/body/main/section[3]/ul/li["+str(t)+"]/a").click()
    element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,f'/html/body/main/section[1]/amp-script/table/tbody/tr[1]/td[1]/a')))

    school = str(driver.find_element(By.XPATH,'/html/body/main/section[1]/amp-script/table/thead/tr/th[1]').text)
    wb = Workbook()
    ws = wb.active
    ws.cell(1,1).value = "科系"
    ws.cell(1,2).value = "檢定標準"
    ws.cell(1,3).value = "採計科目"
    nb = 1
    titles  = driver.find_element(By.XPATH,f'/html/body/main/section[1]/amp-script/table/tbody/tr[{nb}]/td[1]/a')
    levels = driver.find_elements(By.CSS_SELECTOR,"[data-title='檢定標準']") #檢定標準
    takes = driver.find_elements(By.CSS_SELECTOR,"[data-title='採計科目']") #採計科目
    for data in levels:
        title = driver.find_element(By.XPATH,f'/html/body/main/section[1]/amp-script/table/tbody/tr[{nb}]/td[1]').text
        level = driver.find_element(By.XPATH,f'/html/body/main/section[1]/amp-script/table/tbody/tr[{nb}]/td[2]').text
        take = driver.find_element(By.XPATH,f'/html/body/main/section[1]/amp-script/table/tbody/tr[{nb}]/td[3]').text
        # print(title)
        # print(level)
        # print(take)
        ws.cell(nb+1,1).value = title
        ws.cell(nb+1,2).value = level
        ws.cell(nb+1,3).value = take
        nb += 1
    

    wb.save("C:/Users/steve/Desktop/學校/大三下/學習分析工具實務應用/README.md/期末專案/rawdata/"+school+".xlsx")
    driver.back()
driver.close()








