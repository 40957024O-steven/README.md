from openpyxl import load_workbook
import glob

path = "C:/Users/steve/Desktop/學校/大三下/學習分析工具實務應用/README.md/期末專案/rawdata/*.xlsx"
result  =glob.glob(path)
for f in result:
    print(f)
# f = 'C:/Users/steve/Desktop/學校/大三下/學習分析工具實務應用/README.md/期末專案/rawdata\世新大學.xlsx'

    wb = load_workbook(f)
    ws = wb['Sheet']

    rowmax = ws.max_row #行
    # print(rowmax)

    ws.cell(1,4).value = "國文"
    ws.cell(1,5).value = "英文"
    ws.cell(1,6).value = "數A"
    ws.cell(1,7).value = "數B"
    ws.cell(1,8).value = "社會"
    ws.cell(1,9).value = "自然"
    ws.cell(1,10).value = "英聽"
    ws.cell(1,11).value = "數甲"
    ws.cell(1,12).value = "物理"
    ws.cell(1,13).value = "化學"
    ws.cell(1,14).value = "生物"
    ws.cell(1,15).value = "歷史"
    ws.cell(1,16).value = "地理"
    ws.cell(1,17).value = "公民"


    for i in range(2,rowmax+1):
        # print(i)
        tg = ws.cell(i,3).value
        # print(tg)
        vs = tg.split('、')
        for v in vs:
            # print(v)
            if v[:2] == '國文':
                ws.cell(i,4).value = v[3:7]
            elif  v[:2] =='英文':
                ws.cell(i,5).value = v[3:7]
            elif  v[:2] =='數A':
                ws.cell(i,6).value = v[3:7]
            elif  v[:2] =='數B':
                ws.cell(i,7).value = v[3:7]
            elif  v[:2] =='社會':
                ws.cell(i,8).value = v[3:7]
            elif  v[:2] =='自然':
                ws.cell(i,9).value = v[3:7]
            elif  v[:2] =='英聽':
                ws.cell(i,10).value = v[3:7]
            elif  v[:2] =='數甲':
                ws.cell(i,11).value = v[3:7]
            elif  v[:2] =='物理':
                ws.cell(i,12).value = v[3:7]
            elif  v[:2] =='化學':
                ws.cell(i,13).value = v[3:7]
            elif  v[:2] =='生物':
                ws.cell(i,14).value = v[3:7]
            elif  v[:2] =='歷史':
                ws.cell(i,15).value = v[3:7]
            elif  v[:2] =='地理':
                ws.cell(i,16).value = v[3:7]
            elif  v[:2] =='公民':
                ws.cell(i,17).value = v[3:7]

    wb.save(f)